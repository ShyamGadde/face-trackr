import multiprocessing
import os
import queue

import cv2
from cvzone import cornerRect
import face_recognition
import numpy as np
from datetime import datetime

IMG_BACKGROUND = cv2.imread("assets/background.png")


def detect_faces(faces_queue, exit_flag):
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)
    cap.set(4, 480)
    counter = 0

    while True:
        ret, frame = cap.read()

        if not ret:
            break

        IMG_BACKGROUND[162 : 162 + 480, 55 : 55 + 640] = frame
        face_locations = face_recognition.face_locations(frame)

        counter += 1
        if counter % 10 == 0:
            # Only send every tenth frame to the other process
            faces_queue.put((frame, face_locations))

        for top, right, bottom, left in face_locations:
            cornerRect(
                IMG_BACKGROUND, (55 + left, 162 + top, right - left, bottom - top), rt=0
            )

        cv2.imshow("Camera", IMG_BACKGROUND)
        cv2.waitKey(1)

        if cv2.getWindowProperty("Camera", cv2.WND_PROP_VISIBLE) < 1:
            exit_flag.value = 1
            break

    cap.release()
    cv2.destroyAllWindows()


def cache_database(path):
    face_encodings = []
    names = []
    roll = []
    for filename in os.listdir(path):
        # Fix so that it doesn't read hidden files
        if not filename.startswith("."):
            img = cv2.imread(os.path.join(path, filename), cv2.COLOR_BGR2RGB)
            face_encodings.append(face_recognition.face_encodings(img)[0])
            names.append(" ".join(filename.split("_")[1:]).split(".")[0].title())
            roll.append(int(filename.split("_")[0]))
    return face_encodings, names, roll


def process_frame(faces_queue, exit_flag, attendees):
    known_face_encodings, known_face_names, known_face_roll = cache_database(
        "Student_DB"
    )
    counter = 0
    while True:
        try:
            frame, face_locations = faces_queue.get(timeout=1)
        except queue.Empty:
            if exit_flag.value:
                break
            else:
                continue
        face_encodings = face_recognition.face_encodings(frame, face_locations)

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(
                known_face_encodings, face_encoding
            )

            face_distances = face_recognition.face_distance(
                known_face_encodings, face_encoding
            )
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                roll = known_face_roll[best_match_index]

                if roll not in attendees:
                    attendees[roll] = (name, datetime.now().strftime("%I:%M %p"))


def create_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Roll Number", "Name", "Time"])
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 20
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 25
    worksheet.column_dimensions["C"].width = 12

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name="Times New Roman", bold=True, size=14)
    bd = Side(style="thin", color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.fill = PatternFill("solid", fgColor="FFFF00")
    highlight.alignment = Alignment(horizontal="center", vertical="center")

    worksheet["A1"].style = highlight
    worksheet["B1"].style = highlight
    worksheet["C1"].style = highlight
    return workbook


def generate_attendance_report(attendees):
    workbook = create_workbook()
    worksheet = workbook.active

    for roll_num in attendees:
        worksheet.append([roll_num, attendees[roll_num][0], attendees[roll_num][1]])

    workbook.save(f'attendance-records/{datetime.now().strftime("%d-%m-%y (%I.%M-%p)")}.xlsx')


if __name__ == "__main__":
    faces_queue = multiprocessing.Queue()
    exit_flag = multiprocessing.Value("i", 0)
    attendees = multiprocessing.Manager().dict()

    detect_faces_process = multiprocessing.Process(
        target=detect_faces,
        args=(
            faces_queue,
            exit_flag,
        ),
    )
    detect_faces_process.start()
    process_frame_process = multiprocessing.Process(
        target=process_frame,
        args=(
            faces_queue,
            exit_flag,
            attendees,
        ),
    )
    process_frame_process.start()

    detect_faces_process.join()
    process_frame_process.join()

    generate_attendance_report(attendees)
