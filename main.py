from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import face_recognition
import cv2 as cv
import numpy as np
import os
from datetime import datetime
import pickle
from qt_material import apply_stylesheet


class UiMainWindow(object):
    def __init__(self):
        self.action_update_database = None
        self.statusbar = None
        self.menuManage = None
        self.menubar = None
        self.attendance_records = None
        self.take_attendance = None
        self.background = None
        self.central_widget = None

    def setup_ui(self, main_window):
        main_window.setObjectName("Attendance Monitoring System")
        main_window.resize(1280, 768)
        main_window.setMaximumSize(QtCore.QSize(1280, 768))
        main_window.setWindowIcon(QtGui.QIcon('assets/ams.ico'))
        apply_stylesheet(main_window, theme='dark_blue.xml')
        self.central_widget = QtWidgets.QWidget(main_window)
        self.central_widget.setObjectName("central_widget")
        self.background = QtWidgets.QLabel(self.central_widget)
        self.background.setGeometry(QtCore.QRect(0, 0, 1281, 741))
        self.background.setAutoFillBackground(True)
        self.background.setText("")
        self.background.setPixmap(QtGui.QPixmap("assets/gui_background.jpeg"))
        self.background.setScaledContents(True)
        self.background.setIndent(0)
        self.background.setObjectName("background")
        self.take_attendance = QtWidgets.QPushButton(self.central_widget)
        self.take_attendance.setGeometry(840, 440, 319, 65)
        size_policy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        size_policy.setHorizontalStretch(0)
        size_policy.setVerticalStretch(0)
        size_policy.setHeightForWidth(self.take_attendance.sizePolicy().hasHeightForWidth())
        self.take_attendance.setSizePolicy(size_policy)
        font = QtGui.QFont()
        font.setFamily("Roboto Slab")
        font.setPointSize(14)
        self.take_attendance.setFont(font)
        self.take_attendance.setAutoFillBackground(False)
        self.take_attendance.setIconSize(QtCore.QSize(30, 30))
        self.take_attendance.setCheckable(False)
        self.take_attendance.setFlat(True)
        self.take_attendance.setStyleSheet("border : 2px solid white; "
                                           "border-radius : 32px; "
                                           "color : white; "
                                           "font : 14pt Roboto Slab")
        self.take_attendance.clicked.connect(lambda: detect_faces(0))
        self.take_attendance.setObjectName("take_attendance")
        self.attendance_records = QtWidgets.QPushButton(self.central_widget)
        self.attendance_records.setGeometry(840, 540, 319, 65)
        size_policy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        size_policy.setHorizontalStretch(0)
        size_policy.setVerticalStretch(0)
        size_policy.setHeightForWidth(self.attendance_records.sizePolicy().hasHeightForWidth())
        self.attendance_records.setSizePolicy(size_policy)
        font = QtGui.QFont()
        font.setFamily("Roboto Slab")
        font.setPointSize(14)
        self.attendance_records.setFont(font)
        self.attendance_records.setAutoFillBackground(False)
        self.attendance_records.setIconSize(QtCore.QSize(30, 30))
        self.attendance_records.setCheckable(False)
        self.attendance_records.setFlat(True)
        self.attendance_records.setStyleSheet("border : 2px solid white; "
                                              "border-radius : 32px; "
                                              "color : white; "
                                              "font : 14pt Roboto Slab")
        self.attendance_records.clicked.connect(view_attendance_records)
        self.attendance_records.setObjectName("attendance_records")
        main_window.setCentralWidget(self.central_widget)
        self.menubar = QtWidgets.QMenuBar(main_window)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1280, 26))
        self.menubar.setAutoFillBackground(True)
        self.menubar.setDefaultUp(False)
        self.menubar.setNativeMenuBar(True)
        self.menubar.setObjectName("menubar")
        self.menuManage = QtWidgets.QMenu(self.menubar)
        font = QtGui.QFont()
        font.setFamily("Roboto Slab")
        self.menuManage.setFont(font)
        self.menuManage.setAutoFillBackground(True)
        self.menuManage.setTearOffEnabled(False)
        self.menuManage.setSeparatorsCollapsible(True)
        self.menuManage.setToolTipsVisible(True)
        self.menuManage.setObjectName("menuManage")
        main_window.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(main_window)
        self.statusbar.setObjectName("statusbar")
        main_window.setStatusBar(self.statusbar)
        self.action_update_database = QtWidgets.QAction(main_window)
        self.action_update_database.setCheckable(False)
        font = QtGui.QFont()
        font.setFamily("Roboto Slab")
        self.action_update_database.setFont(font)
        self.action_update_database.triggered.connect(update_database)
        self.action_update_database.setObjectName("action_update_database")
        self.menuManage.addAction(self.action_update_database)
        self.menubar.addAction(self.menuManage.menuAction())

        self.retranslate_ui(main_window)
        QtCore.QMetaObject.connectSlotsByName(main_window)

    def retranslate_ui(self, main_window):
        _translate = QtCore.QCoreApplication.translate
        main_window.setWindowTitle(_translate("MainWindow", "Attendance Monitoring System"))
        self.take_attendance.setText(_translate("MainWindow", "Mark Attendance"))
        self.attendance_records.setText(_translate("MainWindow", "Attendance Records"))
        self.menuManage.setTitle(_translate("MainWindow", "Manage"))
        self.action_update_database.setText(_translate("MainWindow", "Update Database"))
        self.action_update_database.setShortcut(_translate("MainWindow", "Ctrl+R"))


# Create arrays of known face encodings and their names and roll numbers
known_face_encodings, known_face_names, known_roll_numbers = [], [], []


def get_face_encodings(file):
    """
    Loads an image file into a numpy array and returns the 128-dimension face encoding for the face in the image

    :param file: Image file name or file object to load
    :return: A list of 128-dimensional face encodings (one for each face in the image)
    """

    return face_recognition.face_encodings(cv.cvtColor(cv.cvtColor(face_recognition.load_image_file(file),
                                                                   cv.COLOR_BGR2GRAY), cv.COLOR_GRAY2RGB))[0]


def create_database(folder):
    """
    Returns a tuple containing three lists, each containing the face encodings, names and roll numbers of each student.

    Given that all the files in the directory are stored in the format roll_first-name_last-name.ext, this function
    returns a tuple containing three lists, each containing the face encodings, names and roll numbers of each student
    all at the exact same index value of each list

    :param folder: The directory with all the images to encode
    :return: A tuple of three lists each containing the face encodings, names and roll numbers, all at the exact same
    index value of each list for a respective image.
    """

    face_encodings = []
    names = []
    roll = []
    for filename in os.listdir(folder):
        if not filename.startswith('.'):
            face_encodings.append(get_face_encodings(os.path.join(folder, filename)))
            names.append(' '.join(filename.split('_')[1:]).split('.')[0].title())
            roll.append(int(filename.split('_')[0]))
    return face_encodings, names, roll


def mark_attendance(roll, name, excel):
    """
    A simple function to mark the attendance given that an entry already doesn't exist

    :param roll: The roll number of the person whose face was recognized
    :param name: The name of the identified person
    :param excel: The Excel file in which to make the entries
    :return: None
    """

    worksheet = excel.active
    # Create a list existing roll numbers in the worksheet.
    roll_numbers_list = list(*worksheet.iter_cols(max_col=1, values_only=True))[1:]
    if roll not in roll_numbers_list:
        worksheet.append([roll, name, datetime.now().strftime("%I:%M:%S %p")])


def create_workbook():
    """
    A simple to create an Excel workbook

    :return: An Excel workbook object
    """

    from openpyxl import Workbook
    from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(['Roll Number', 'Name', 'Time'])
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 20
    worksheet.column_dimensions['A'].width = 14
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 12

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Times New Roman', bold=True, size=12)
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.fill = PatternFill('solid', fgColor='FFFF00')
    highlight.alignment = Alignment(horizontal="center", vertical="center")

    worksheet["A1"].style = highlight
    worksheet["B1"].style = highlight
    worksheet["C1"].style = highlight
    return workbook


def detect_faces(cam_id=0):
    """
    Detects faces in a given video stream

    :param cam_id: ID of the camera device. Set to zero by default
    :return: None
    """

    load_database()

    workbook = create_workbook()

    # Get a reference to webcam
    video_capture = cv.VideoCapture(cam_id)

    while True:
        # Grab a single frame of video
        _, frame = video_capture.read()

        gray_frame = cv.cvtColor(cv.cvtColor(frame, cv.COLOR_BGR2GRAY), cv.COLOR_GRAY2RGB)

        face_locations = face_recognition.face_locations(gray_frame)
        face_encodings = face_recognition.face_encodings(gray_frame, face_locations)

        # Loop through each face in this frame of video
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            name = "Unknown"
            roll = ""

            # Use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                roll = known_roll_numbers[best_match_index]
                mark_attendance(roll, name, workbook)

            # Draw a bounding box around the face
            cv.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            # Draw a label with a roll number and name below the face
            label_box = frame.copy()
            cv.rectangle(label_box, (left, bottom - 25), (right, bottom), (0, 255, 0), cv.FILLED)
            alpha = 0.3
            frame = cv.addWeighted(label_box, alpha, frame, 1 - alpha, 0)
            # Dynamically change the font size according to the bounding box
            font_size = (right - left) / 372
            cv.putText(frame, f'{roll} {name}', (left, bottom - 6), 0, font_size, (255, 255, 255), 1)
        cv.imshow('Face Recognition', cv.resize(frame, (0, 0), fx=1.6, fy=1.6))

        # Hit 'q' on the keyboard to quit!
        if cv.waitKey(1) & 0xFF == ord('q'):
            workbook.save(f'Attendance_records/{datetime.now().strftime("%d-%m-%y (%I.%M-%p)")}.xlsx')
            break

    # Release handle to the webcam
    video_capture.release()
    cv.destroyAllWindows()


def update_database():
    """
    Update the database.pkl file, incase any changes in the 'Student_DB'.

    :return: None
    """
    face_encodings, names, roll_numbers = create_database("Student_DB")

    open('database.pkl', 'w').close()
    db = open('database.pkl', 'wb')
    pickle.dump(face_encodings, db)
    pickle.dump(names, db)
    pickle.dump(roll_numbers, db)
    db.close()

    msg = QMessageBox()
    msg.setWindowTitle("Update")
    msg.setWindowIcon(QtGui.QIcon('assets/ams.ico'))
    msg.setText("Database updated successfully!")
    msg.setIcon(QMessageBox.Information)

    _ = msg.exec_()


def view_attendance_records():
    """
    A simple function to open the folder containing the attendance records.

    :return: None
    """

    os.startfile('Attendance_records')


def load_database():
    """
    A function to load the face encodings, names and roll numbers from the database into the program.

    :return: None
    """

    global known_face_encodings, known_face_names, known_roll_numbers

    database = open('database.pkl', 'rb')
    known_face_encodings = pickle.load(database)
    known_face_names = pickle.load(database)
    known_roll_numbers = pickle.load(database)
    database.close()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = UiMainWindow()
    ui.setup_ui(MainWindow)
    MainWindow.show()
    if not os.path.exists('database.pkl'):
        update_database()
    sys.exit(app.exec_())


# <a href="https://www.flaticon.com/free-icons/face" title="face icons">Face icons created by juicy_fish - Flaticon</a>
# https://www.vecteezy.com/vector-art/1227414-abstract-modern-gradient-flowing-geometric-background
