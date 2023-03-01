from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from typing import Dict

def create_workbook():
    """
    It creates a workbook with a worksheet, adds a header row, and sets the style of the header row
    :return: A workbook object.
    """
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Student ID", "Name", "Time"])
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 20
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 15

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name="Times New Roman", bold=True, size=14)
    bd = Side(style="thin", color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.fill = PatternFill("solid", fgColor="FFFF00")
    highlight.alignment = Alignment(horizontal="center", vertical="center")

    worksheet["A1"].style = highlight
    worksheet["B1"].style = highlight
    worksheet["C1"].style = highlight
    return workbook


def generate_attendance_report(attendees):
    """
    It takes a dictionary of attendees and creates an excel file with the attendance records
    
    :param attendees: A dictionary containing the roll numbers of the students who attended the class as
    keys and a list with their names and corresponding attendance timestamps in that order as values
    """
    workbook = create_workbook()
    worksheet = workbook.active

    for roll_num in attendees:
        worksheet.append([roll_num, attendees[roll_num][0], attendees[roll_num][1]])

    workbook.save(f'attendance-records/{datetime.now().strftime("%d-%m-%y (%I.%M-%p)")}.xlsx')
